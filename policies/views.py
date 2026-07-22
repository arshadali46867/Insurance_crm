from django.shortcuts import render

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
import openpyxl
from django.http import HttpResponse

from datetime import date, timedelta

from .models import Policy
from .serializers import PolicySerializer
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse


class PolicyViewSet(viewsets.ModelViewSet):

    queryset = Policy.objects.all().order_by("-created_at")

    serializer_class = PolicySerializer

    permission_classes = [
        IsAuthenticated
    ]


    @action(
        detail=False,
        methods=['get']
    )
    def renewals(self, request):

        today = date.today()

        upcoming_date = today + timedelta(days=30)


        policies = Policy.objects.filter(
            expiry_date__range=[
                today,
                upcoming_date
            ]
        )


        data = []

        for policy in policies:

            data.append({
                "customer": policy.customer.name,
                "policy_number": policy.policy_number,
                "insurance_type": policy.insurance_type,
                "expiry_date": policy.expiry_date,
                "days_left": (
                    policy.expiry_date - today
                ).days
            })


        return Response(data)
    
def export_policies_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Policies"

    headers = [
        "Policy Number",
        "Customer",
        "Insurance Type",
        "Company",
        "Premium",
        "Start Date",
        "Expiry Date"
    ]

    for col_num, header in enumerate(headers, 1):

        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    policies = Policy.objects.select_related(
        "customer"
    )

    row_num = 2

    for policy in policies:

        sheet.cell(
            row=row_num,
            column=1
        ).value = policy.policy_number

        sheet.cell(
            row=row_num,
            column=2
        ).value = policy.customer.name

        sheet.cell(
            row=row_num,
            column=3
        ).value = policy.insurance_type

        sheet.cell(
            row=row_num,
            column=4
        ).value = policy.insurance_company

        sheet.cell(
            row=row_num,
            column=5
        ).value = float(policy.premium_amount)

        sheet.cell(
            row=row_num,
            column=6
        ).value = str(policy.start_date)

        sheet.cell(
            row=row_num,
            column=7
        ).value = str(policy.expiry_date)

        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="policies.xlsx"'

    workbook.save(response)

    return response    



def export_policies_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="policies.pdf"'

    pdf = SimpleDocTemplate(response)

    data = [
        [
            "Policy No",
            "Customer",
            "Type",
            "Company",
            "Premium",
            "Expiry Date"
        ]
    ]

    policies = Policy.objects.select_related(
        "customer"
    )

    for policy in policies:

        data.append([
            policy.policy_number,
            policy.customer.name,
            policy.insurance_type,
            policy.insurance_company,
            str(policy.premium_amount),
            str(policy.expiry_date)
        ])

    table = Table(data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])
    )

    pdf.build([table])

    return response
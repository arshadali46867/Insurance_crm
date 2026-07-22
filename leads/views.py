from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter

from .models import Lead
from .serializers import LeadSerializer

from customers.models import Customer
from accounts.models import User
import openpyxl

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from activity_logs.models import ActivityLog


class LeadViewSet(viewsets.ModelViewSet):

    queryset = Lead.objects.all()

    serializer_class = LeadSerializer

    permission_classes = [
        IsAuthenticated
    ]

    filter_backends = [
        DjangoFilterBackend,
        SearchFilter
    ]

    filterset_fields = [
        "status"
    ]

    search_fields = [
        "name",
        "mobile",
        "email"
    ]

    def get_queryset(self):

        user = self.request.user

        # Agent sirf apni leads dekhega
        if user.role == "agent":

            return Lead.objects.filter(
                assigned_to=user
            ).order_by("-created_at")

        # Admin aur Manager sab leads dekhenge
        return Lead.objects.all().order_by("-created_at")

    def perform_create(self, serializer):

        lead = serializer.save(
            created_by=self.request.user
        )

        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Created Lead {lead.name}"
        )

    @action(
        detail=True,
        methods=['post']
    )
    def convert(self, request, pk=None):

        lead = self.get_object()

        customer, created = Customer.objects.get_or_create(
            email=lead.email,
            defaults={
                "name": lead.name,
                "mobile": lead.mobile
            }
        )

        lead.status = "converted"
        lead.save()
        ActivityLog.objects.create(
        user=request.user,
        action=f"Converted Lead {lead.name} into Customer"
        )

        return Response({
            "message": "Lead converted successfully",
            "customer_id": customer.id,
            "customer_name": customer.name
        })

    @action(
        detail=True,
        methods=['post']
    )
    def assign(self, request, pk=None):

        lead = self.get_object()

        agent_id = request.data.get("agent_id")

        if not agent_id:
            return Response(
                {
                    "error": "agent_id is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            agent = User.objects.get(
                id=agent_id,
                role="agent"
            )

        except User.DoesNotExist:

            return Response(
                {
                    "error": "Agent not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        lead.assigned_to = agent
        lead.save()
        ActivityLog.objects.create(
        user=request.user,
        action=f"Assigned Lead {lead.name} to {agent.username}"
        )

        return Response({
            "message": "Lead assigned successfully",
            "lead_id": lead.id,
            "agent_id": agent.id,
            "agent": agent.username
        })
    


def export_leads_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Leads"

    headers = [
        "ID",
        "Name",
        "Mobile",
        "Email",
        "Status",
        "Assigned To"
    ]

    for col_num, header in enumerate(headers, 1):

        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    leads = Lead.objects.select_related(
        "assigned_to"
    )

    row_num = 2

    for lead in leads:

        sheet.cell(
            row=row_num,
            column=1
        ).value = lead.id

        sheet.cell(
            row=row_num,
            column=2
        ).value = lead.name

        sheet.cell(
            row=row_num,
            column=3
        ).value = lead.mobile

        sheet.cell(
            row=row_num,
            column=4
        ).value = lead.email

        sheet.cell(
            row=row_num,
            column=5
        ).value = lead.status

        sheet.cell(
            row=row_num,
            column=6
        ).value = (
            lead.assigned_to.username
            if lead.assigned_to
            else ""
        )

        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="leads.xlsx"'

    workbook.save(response)

    return response    


def export_leads_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="leads.pdf"'

    pdf = SimpleDocTemplate(response)

    data = [[
        "ID",
        "Name",
        "Mobile",
        "Email",
        "Status"
    ]]

    leads = Lead.objects.all()

    for lead in leads:

        data.append([
            lead.id,
            lead.name,
            lead.mobile,
            lead.email or "",
            lead.status
        ])

    table = Table(data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )

    pdf.build([table])

    return response
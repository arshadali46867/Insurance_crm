from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Customer
from .serializers import CustomerSerializer
from policies.models import Policy
import openpyxl
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse



class CustomerViewSet(viewsets.ModelViewSet):

    queryset = Customer.objects.all().order_by("-created_at")

    serializer_class = CustomerSerializer

    permission_classes = [
        IsAuthenticated
    ]


    @action(
        detail=True,
        methods=['get']
    )
    def policies(self, request, pk=None):

        customer = self.get_object()

        policies = Policy.objects.filter(
            customer=customer
        )

        data = []

        for policy in policies:
            data.append({
                "policy_number": policy.policy_number,
                "insurance_type": policy.insurance_type,
                "insurance_company": policy.insurance_company,
                "premium_amount": policy.premium_amount,
                "start_date": policy.start_date,
                "expiry_date": policy.expiry_date
            })


        return Response({
            "customer": customer.name,
            "policies": data
        })
    
def export_customers_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Customers"

    headers = [
        "ID",
        "Name",
        "Email",
        "Mobile",
        "Address",
        "Created At"
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    customers = Customer.objects.all()

    row_num = 2

    for customer in customers:

        sheet.cell(row=row_num, column=1).value = customer.id
        sheet.cell(row=row_num, column=2).value = customer.name
        sheet.cell(row=row_num, column=3).value = customer.email
        sheet.cell(row=row_num, column=4).value = customer.mobile
        sheet.cell(row=row_num, column=5).value = customer.address
        sheet.cell(row=row_num, column=6).value = str(customer.created_at)

        row_num += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="customers.xlsx"'
    )

    workbook.save(response)

    return response    

def export_customers_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="customers.pdf"'

    pdf = SimpleDocTemplate(response)

    data = [[
        "ID",
        "Name",
        "Email",
        "Mobile",
        "Address"
    ]]

    customers = Customer.objects.all()

    for customer in customers:

        data.append([
            customer.id,
            customer.name,
            customer.email,
            customer.mobile,
            customer.address or ""
        ])

    table = Table(data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )

    pdf.build([table])

    return response